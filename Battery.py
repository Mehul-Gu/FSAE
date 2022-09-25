import openpyxl
import numpy as np

dfs = openpyxl.load_workbook("Cell Choice Python Calculator.xlsx")
data = dfs.active

values = []
all_values = []

for row in range(1, data.max_row):
    for col in data.iter_cols(0, data.max_column):
        values.append(col[row].value)
    all_values.append(values)
    values = []

maximum_voltage = 420
total_capacity = 6900

i=0
for values in all_values:
    #grabbing and calcing some basic stuff
    name = values[0]
    energy_per_cell = values[1] * values[4]
    number_of_cells = int(total_capacity / energy_per_cell) + 1

    #standardizing discharge based on excel inputs
    if values[2] == None:
        discharge = values[3] * values[1]
    else:
        discharge = values[2]

    #calcing power and checking if we have enough
    power_limit_check = discharge * number_of_cells * values[4] * 2
    if power_limit_check > 80000:
        power_check = True
    else:
        power_check = False

    #iterate through number of cells to make it a very composite number
    counter = 0
    i = 1
    temp = number_of_cells
    new_number = number_of_cells
    check = False
    while check == False:
        if (temp < 6):
            number_of_cells = new_number
            check == True
            break
        elif temp % 2 == 0:
            counter += 1
            temp = temp/2
        elif temp % 3 == 0:
            counter += 1
            temp = temp/3
        else:
            temp = number_of_cells
            temp += i
            new_number = temp
            i += 1
    
    #iterate number of cells/series cells until it is perfectly divisible
    cells_in_series = maximum_voltage / values[5]
    check = False
    while check == False:
        if number_of_cells % cells_in_series != 0:
            cells_in_series -= 1
        else:
            check = True
    
    parallel = number_of_cells / cells_in_series

    #6MJ and 120V segment requirement calcs
    check = False
    i = 1
    energy_per_cell_joules = energy_per_cell * 3600
    while check == False:
        if cells_in_series % i == 0:
            series_in_segment = cells_in_series/i
            total_cells_in_segment = series_in_segment * parallel
            if ((total_cells_in_segment * energy_per_cell_joules) < (6000000)) and ((series_in_segment * values[4]) < 120):
                check = True
                break
            else:
                i += 1
        else:
            i += 1


    if power_check == True:
        print(f"The cell {name} has passed all requirements")
        print(f"The configuration is {number_of_cells} total cells. {cells_in_series} cells in series and times {parallel} parallel")
        print(f"{total_cells_in_segment} cells in each segment with {parallel} parallel and {series_in_segment} in series")
    else:
        print("The cell does not pass the power check")


    